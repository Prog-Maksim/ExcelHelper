import os
import json
import random
from datetime import datetime

import openpyxl
import xlrd
from pathlib import Path
from openpyxl import Workbook
from cryptography.fernet import Fernet
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill


def download_excel_file(file_path: str):
    """
    Загружает файл для его дальнейшей обработки
    :param file_path: строка, путь до файла
    :return: рабочий лист и обькт табличы
    """

    if file_path.split('.')[-1] == 'xls':
        # загружаем файл xls
        wb_xls = xlrd.open_workbook(file_path)

        # создаем новый файл xlsx
        wb_xlsx = Workbook()

        # выбираем активный лист
        ws = wb_xlsx.active

        # копируем данные из xls в xlsx
        for sheet in wb_xls.sheets():
            for row in range(sheet.nrows):
                values = []
                for col in range(sheet.ncols):
                    values.append(sheet.cell(row, col).value)
                ws.append(values)

        # сохраняем файл xlsx
        wb_xlsx.save("temporality.xlsx")

        workbook = load_workbook(filename='temporality.xlsx', data_only=False)
        worksheet = workbook.active
        return worksheet, workbook
    elif file_path.split('.')[-1] == 'xlsx':
        workbook = load_workbook(filename=file_path, data_only=False)
        worksheet = workbook.active
        return worksheet, workbook


def find_dublication(worksheet, index: list) -> None:
    """
    Нахождение повторяющихся значений между колонками Увеличение и Уменьшение
    :param worksheet: рабочий лист
    :param index: список из двух индексов нахождения столбцов
    :return: None
    """
    col_x, col_y = index
    for row in range(14, worksheet.max_row):
        increase_cell = worksheet.cell(row=row, column=col_x)
        for row1 in range(14, worksheet.max_row + 1):
            decrease_cell = worksheet.cell(row=row1, column=col_y)

            # Проверка, есть ли значение ячейки в обоих столбцах
            if increase_cell.value and decrease_cell.value and increase_cell.value == decrease_cell.value:
                # Пометка ячеек серым цветом
                fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
                increase_cell.fill = fill
                decrease_cell.fill = fill


def find_values(worksheet, index: list):
    """
    Поиск всех значеий в столбцах "Увеличение" и "Уменьшение"
    :param worksheet: рабочий лист
    :param index: список из двух индексов нахождения столбцов
    :return: возвращает список значений для двух столбцов
    """
    col_x, col_y = index
    increase_cell = []
    decrease_cell = []
    for row in range(14, worksheet.max_row):
        increase_cell.append(worksheet.cell(row=row, column=col_x).value)
        decrease_cell.append(worksheet.cell(row=row, column=col_y).value)
    increase_cell = [i for i in increase_cell[1:] if i is not None]
    decrease_cell = [i for i in decrease_cell[1:] if i is not None]
    return increase_cell, decrease_cell


def find_pair(numbers: list, nums: int, used_numbers: list):
    """
    Поиск и нахождение суммы двух чисел в столбце равное числу nums
    :param numbers: список чисел
    :param nums: искомое число
    :param used_numbers: список уже найденных чисел
    :return: возвращает результат найденных двух чисел в столбце
    """

    numbers = [num for num in numbers if num not in used_numbers]
    numbers.sort()
    left = 0
    right = len(numbers) - 1
    while left < right:
        current_sum = round(numbers[left] + numbers[right], 2)
        if current_sum == nums:
            return numbers[left], numbers[right]
        elif current_sum < nums:
            left += 1
        else:
            right -= 1


def processing_data(worksheet, index: list):
    """
    Нахождение двух чисел в одном столбце дающие сумму числа равное другому числу
    :param worksheet: рабочий лист
    :param index: индексы столбцов "Увеличение" и "Уменьшение"
    :return: словарь результат работы программы, список чисел в "Увеличение" и "Уменьшение"
    """
    result = dict()
    number, numbers = find_values(worksheet=worksheet, index=index)
    used_numbers = []
    for i in number:
        pair = find_pair(numbers, i, used_numbers)

        if pair:
            try:
                result[i] = [pair[0], pair[1]]
                used_numbers.extend([pair[0], pair[1]])
            except: pass
    return result, number, numbers


def processing_data_1(worksheet, index: list):
    """
    Нахождение двух чисел в одном столбце дающие сумму числа равное другому числу
    (проверка тех же совпадений, но в противоположных столбцах)
    :param worksheet: рабочий лист
    :param index: индексы столбцов "Увеличение" и "Уменьшение"
    :return: словарь результат работы программы, список чисел в "Увеличение" и "Уменьшение"
    """
    result = dict()
    number, numbers = find_values(worksheet=worksheet, index=index)
    used_numbers = []
    for i in numbers:
        pair = find_pair(number, i, used_numbers)

        if pair:
            try:
                result[i] = [pair[0], pair[1]]
                used_numbers.extend([pair[0], pair[1]])
            except: pass
    return result, number, numbers


def update_table(worksheet, values: dict, index: list) -> None:
    """
    Результат выполнения функций processing_data и processing_data_1 помечаем на таблице серым цветом
    :param worksheet: рабочий лист
    :param values: словарь значений из двух чисел в сумме дающие число -> ключ
    :param index: индексы столбцов "Увеличение" и "Уменьшение"
    :return: None
    """
    col_x, col_y = index
    for row in range(14, worksheet.max_row):
        increase_cell = worksheet.cell(row=row, column=col_x)
        decrease_cell = worksheet.cell(row=row, column=col_y)
        fill = PatternFill(start_color='BFBEC3', end_color='BFBEC3', fill_type='solid')

        for i in values:
            if i == increase_cell.value:
                increase_cell.fill = fill
        for i in values.values():
            if i[0] == decrease_cell.value:
                decrease_cell.fill = fill
            elif i[1] == decrease_cell.value:
                decrease_cell.fill = fill


def update_table_1(worksheet, values, index: list) -> None:
    """
    Результат выполнения функций processing_data и processing_data_1 помечаем на таблице серым цветом
    (функция для противоположных столбцов)
    :param worksheet: рабочий лист
    :param values: словарь значений из двух чисел в сумме дающие число -> ключ
    :param index: индексы столбцов "Увеличение" и "Уменьшение"
    :return: None
    """
    col_x, col_y = index
    for row in range(14, worksheet.max_row):
        increase_cell = worksheet.cell(row=row, column=col_x)
        decrease_cell = worksheet.cell(row=row, column=col_y)
        fill = PatternFill(start_color='BFBEC3', end_color='BFBEC3', fill_type='solid')

        for i in values:
            if i == decrease_cell.value:
                decrease_cell.fill = fill
        for i in values.values():
            if i[0] == increase_cell.value:
                increase_cell.fill = fill
            elif i[1] == increase_cell.value:
                increase_cell.fill = fill


def security_table(worksheets, workbook: openpyxl.Workbook) -> None:
    """
    Устанавливает различные блокировки excel файлу
    :param worksheets: рабочий лист
    :param workbook: обьект таблицы
    :return: None
    """
    with open("json files/security_table.json", 'r', encoding='utf-8') as file:
        data: dict = dict(json.load(file))
        protection = data['state']
        password = data['password'].replace("b'", "").replace("'", "")
        key = "A9bw4erG2-CzbVGE7My4fv6jwVIpEJVdrKqg34ccGkw="
        f = Fernet(key)
        password = str(f.decrypt(password)).replace("b'", "").replace("'", "")

        if protection:
            security: dict = data.get("types_security")
            if security.get("structure_change"):
                workbook.security.workbookPassword = password
                workbook.security.lockStructure = True
            if security.get("history_change"):
                workbook.security.workbookPassword = password
                workbook.security.revisionsPassword = password
            if security.get("prohibion_change"):
                worksheets.protection.password = password


def save_file_format(workbook, worksheets, file_name=None, path=False):
    """
    Сохраняем обработанный excel файл
    :param workbook: обьект таблицы
    :param worksheets: рабочий лист
    :param file_name: имя файла для сохранения
    :param path: путь до сохранения файла
    :return: возвращает результат bool и строку: путь до файла или текст ошибки
    """

    if os.path.exists('json files/table.json'):
        with open('json files/table.json', 'r', encoding='utf-8') as file:
            data = dict(json.load(file))
    else:
        with open('json files/table.json', 'w', encoding='utf-8') as file:
            json.dump({}, file, indent=4, ensure_ascii=False)
            data = dict()

    if os.path.exists('json files/security_table.json'):
        security_table(worksheets=worksheets, workbook=workbook)

    try:
        os.remove('temporality.xlsx')
        if path:
            file_name = f'Таблица №{random.randint(1111, 9999)}.xlsx'
            if not Path('Обработанные excel файлы').is_dir():
                os.mkdir('Обработанные excel файлы')
            workbook.save(filename='Обработанные excel файлы/' + file_name)

            with open('json files/table.json', 'w', encoding='utf-8') as file:
                path_file = os.path.abspath('Обработанные excel файлы')
                data[file_name] = {
                    "file_path": path_file + '\\' + file_name,
                    "data": str(datetime.now().date()),
                    "time": str(datetime.now().time())[0: 5],
                    "state": True
                }
                json.dump(data, file, indent=4, ensure_ascii=False)

            return True, f'Обработанные excel файлы/{file_name}'
        else:
            workbook.save(filename=file_name)

            with open('json files/table.json', 'w', encoding='utf-8') as file:
                data[os.path.basename(file_name)] = {
                    "file_path": file_name,
                    "data": str(datetime.now().date()),
                    "time": str(datetime.now().time())[0: 5],
                    "state": True
                }
                json.dump(data, file, ensure_ascii=False, indent=4)

            return True, file_name
    except PermissionError:
        return False, f'Произошла ошибка: файл "{file_name}" уже открыт. \nЧтобы решить эту проблему, закройте файл'
    except Exception as ex:
        return False, str(ex)
