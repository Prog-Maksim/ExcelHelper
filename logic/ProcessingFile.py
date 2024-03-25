import os
from datetime import datetime

import openpyxl
import xlrd
import tkinter as tk
import customtkinter as ctk
from pathlib import Path
from openpyxl import Workbook
from cryptography.fernet import Fernet
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Protection
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

from UI import complite_processing_file

# TODO Убрать заголовки вврху файла
# TODO Переименовывать название листа

class processing_file:
    def __init__(self, file_path: Path, save_path: Path = None, security: list = None, password: str = None):
        self.file_path = file_path
        self.save_path = save_path
        self.security = security
        self.password = password

        self.all_values = 0
        self.find_same_values = 0

    @staticmethod
    def get_file_size_mb(file_path: Path) -> float:
        file_size = os.path.getsize(file_path)
        file_size_mb = file_size / 1024
        return round(file_size_mb, 2)

    def start(self) -> dict:
        try:
            if self.file_path.suffix == '.xlsx':
                self.workbook = load_workbook(filename=self.file_path, data_only=False)
                self.worksheet = self.workbook.active
            elif self.file_path.suffix == '.xls':
                df = pd.read_excel(self.file_path)
                df.to_excel('temporality.xlsx', index=False)

                self.workbook = load_workbook(filename='temporality.xlsx', data_only=False)
                self.worksheet = self.workbook.active
            elif self.file_path.suffix == '.csv':
                df = pd.read_csv(self.file_path)
                df.to_excel('temporality.xlsx', index=False)

                self.workbook = load_workbook(filename='temporality.xlsx', data_only=False)
                self.worksheet = self.workbook.active

            self.__find_dublication()
            a, _, _ = self.processing_data()
            self.update_table(values=a)
            a, _, _ = self.processing_data_1()
            self.update_table_1(values=a)

            self.security_table(self.password)
            result = self.save_file_format()

            if result:
                result = {
                    "success": True,
                    "data": {
                        "Information": {
                            "all-processing": self.all_values,
                            "same-values": self.find_same_values
                        },
                        "Security": True if self.security else False,
                        "ExcelHelper": {
                            "Name": self.save_path.name,
                            "Weight": self.get_file_size_mb(self.save_path),
                            "Datetime": datetime.now().date()
                        },
                        "Files": {
                            "Path": os.path.abspath(self.save_path)
                        },
                        "Browser": False
                    },
                    "Error": ""
                }
                return result

        except FileNotFoundError as error:
            return {"success": False, "Data": {}, "Error": f"Ошибка сохранения временного файла: {error}"}
        except PermissionError:
            return {"success": False, "Data": {}, "Error": "Сохраняемый файл открыт, закройте его!"}
        except Exception as error:
            print("Что-то страшное")
            return {"success": False, "Data": {}, "Error": f"Ошибка: {error}"}
        finally:
            if os.path.exists("temporality.xlsx"):
                os.remove("temporality.xlsx")

    def search_index(self) -> tuple[int, int]:
        col_x = 0
        col_y = 0
        for i in range(1, self.worksheet.max_row + 1):
            for j in range(1, self.worksheet.max_column + 1):
                index = self.worksheet.cell(row=i, column=j).value
                if index == "Увеличение":
                    col_x = j
                elif index == "Уменьшение":
                    col_y = j

        return col_x, col_y

    def __find_dublication(self) -> None:
        col_x, col_y = self.search_index()

        for row in range(14, self.worksheet.max_row):
            increase_cell = self.worksheet.cell(row=row, column=col_x)

            for row1 in range(14, self.worksheet.max_row + 1):
                decrease_cell = self.worksheet.cell(row=row1, column=col_y)

                if increase_cell.value and decrease_cell.value and increase_cell.value == decrease_cell.value:
                    self.find_same_values += 1
                    fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
                    increase_cell.fill = fill
                    decrease_cell.fill = fill

    def __find_values(self) -> tuple[list, list]:
        """
        Поиск всех значеий в столбцах "Увеличение" и "Уменьшение"
        :return: возвращает список значений для двух столбцов
        """
        col_x, col_y = self.search_index()
        increase_cell = []
        decrease_cell = []
        for row in range(14, self.worksheet.max_row):
            increase_cell.append(self.worksheet.cell(row=row, column=col_x).value)
            decrease_cell.append(self.worksheet.cell(row=row, column=col_y).value)
        increase_cell = [i for i in increase_cell[1:] if i is not None]
        decrease_cell = [i for i in decrease_cell[1:] if i is not None]
        return increase_cell, decrease_cell

    def find_pair(self, numbers: list, nums: int, used_numbers: list) -> tuple[int, int]:
        """
        Поиск и нахождение суммы двух чисел в столбце равное числу nums
        :param numbers: список чисел
        :param nums: искомое число
        :param used_numbers: список уже найденных чисел
        :return: возвращает результат найденных двух чисел в столбце
        """

        numbers = [num for num in numbers if num not in used_numbers]
        numbers.sort()
        left = 0
        right = len(numbers) - 1
        while left < right:
            current_sum = round(numbers[left] + numbers[right], 2)
            if current_sum == nums:
                self.all_values += 1
                return numbers[left], numbers[right]
            elif current_sum < nums:
                left += 1
            else:
                right -= 1

    def processing_data(self) -> tuple:
        """
        Нахождение двух чисел в одном столбце дающие сумму числа равное другому числу
        :param worksheet: рабочий лист
        :param index: индексы столбцов "Увеличение" и "Уменьшение"
        :return: словарь результат работы программы, список чисел в "Увеличение" и "Уменьшение"
        """
        result = dict()
        number, numbers = self.__find_values()
        used_numbers = []
        for i in number:
            pair = self.find_pair(numbers, i, used_numbers)

            if pair:
                try:
                    result[i] = [pair[0], pair[1]]
                    used_numbers.extend([pair[0], pair[1]])
                except:
                    pass
        return result, number, numbers

    def processing_data_1(self) -> tuple:
        """
        Нахождение двух чисел в одном столбце дающие сумму числа равное другому числу
        (проверка тех же совпадений, но в противоположных столбцах)
        :param worksheet: рабочий лист
        :param index: индексы столбцов "Увеличение" и "Уменьшение"
        :return: словарь результат работы программы, список чисел в "Увеличение" и "Уменьшение"
        """
        result = dict()
        number, numbers = self.__find_values()
        used_numbers = []
        for i in numbers:
            pair = self.find_pair(number, i, used_numbers)

            if pair:
                try:
                    result[i] = [pair[0], pair[1]]
                    used_numbers.extend([pair[0], pair[1]])
                except:
                    pass
        return result, number, numbers

    def update_table(self, values: dict) -> None:
        """
        Результат выполнения функций processing_data и processing_data_1 помечаем на таблице серым цветом
        :param worksheet: рабочий лист
        :param values: словарь значений из двух чисел в сумме дающие число -> ключ
        :param index: индексы столбцов "Увеличение" и "Уменьшение"
        :return: None.
        """
        col_x, col_y = self.search_index()
        for row in range(14, self.worksheet.max_row):
            increase_cell = self.worksheet.cell(row=row, column=col_x)
            decrease_cell = self.worksheet.cell(row=row, column=col_y)
            fill = PatternFill(start_color='BFBEC3', end_color='BFBEC3', fill_type='solid')

            for i in values:
                if i == increase_cell.value:
                    increase_cell.fill = fill
            for i in values.values():
                if i[0] == decrease_cell.value:
                    decrease_cell.fill = fill
                elif i[1] == decrease_cell.value:
                    decrease_cell.fill = fill

    def update_table_1(self, values, ) -> None:
        """
        Результат выполнения функций processing_data и processing_data_1 помечаем на таблице серым цветом
        (функция для противоположных столбцов)
        :param worksheet: рабочий лист
        :param values: словарь значений из двух чисел в сумме дающие число -> ключ
        :param index: индексы столбцов "Увеличение" и "Уменьшение"
        :return: None.
        """
        col_x, col_y = self.search_index()
        for row in range(14, self.worksheet.max_row):
            increase_cell = self.worksheet.cell(row=row, column=col_x)
            decrease_cell = self.worksheet.cell(row=row, column=col_y)
            fill = PatternFill(start_color='BFBEC3', end_color='BFBEC3', fill_type='solid')

            for i in values:
                if i == decrease_cell.value:
                    decrease_cell.fill = fill
            for i in values.values():
                if i[0] == increase_cell.value:
                    increase_cell.fill = fill
                elif i[1] == increase_cell.value:
                    increase_cell.fill = fill

    def security_table(self, password: str) -> None:
        if not self.security:
            return

        print(password)
        if "Защита изменения структуры таблицы" in self.security:
            self.workbook.security.workbookPassword = password
            self.workbook.security.lockStructure = True
        if "Защита отслеживания истории изменения" in self.security:
            self.workbook.security.workbookPassword = password
            self.workbook.security.revisionsPassword = password
        if "Защита листа от изменения" in self.security:
            self.worksheet.protection.password = password

    def save_file_format(self) -> bool:
        self.workbook.save(filename=self.save_path)
        return True
